from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def format_excel_style(excel_path: str):
    """
    调整Excel文件的列宽和换行样式
    
    参数
    ------
    excel_path : str
        Excel文件路径
    """
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # 需要设置列宽和换行的列（列宽增大 + 数据换行）
        wrap_columns = {
            "商品链接": 30,
            "产品标题": 40,
            "核心词周期": 40,
        }
        
        # 只需要增大列宽的列
        width_only_columns = {
            "asin": 15,
            "主题": 15,
            "上架时间": 15,
        }
        
        # 创建自动换行的对齐方式
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        # 找到目标列的位置并设置宽度和换行
        header_row = 1
        for col_idx, cell in enumerate(ws[header_row], 1):
            column_name = cell.value
            
            # 处理需要换行的列
            if column_name in wrap_columns:
                col_letter = get_column_letter(col_idx)
                width = wrap_columns[column_name]
                
                # 设置列宽
                ws.column_dimensions[col_letter].width = width
                
                # 为所有数据行设置自动换行
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row_idx, col_idx)
                    cell.alignment = wrap_alignment
            
            # 处理只需要增大列宽的列
            elif column_name in width_only_columns:
                col_letter = get_column_letter(col_idx)
                width = width_only_columns[column_name]
                
                # 设置列宽
                ws.column_dimensions[col_letter].width = width
        
        wb.save(excel_path)
        print(f'已调整Excel文件样式: {excel_path}')
        
    except Exception as e:
        print(f'调整Excel文件样式时出错: {e}')
        import traceback
        traceback.print_exc()

