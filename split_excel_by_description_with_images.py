from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Fill, PatternFill, Border, Side
import os


def excel_colwidth_to_pixels(width):
    if width is None:
        return 100
    return int(width * 7)


def excel_rowheight_to_pixels(height):
    if height is None:
        return 80
    return int(height * 1.33)


def extract_main_category(conclusion_value):
    """
    从开发结论值中提取主要类别
    例如："开发-价格上升" -> "开发"
         "待定-价格上升" -> "待定"
    """
    if not conclusion_value:
        return "未分类"
    conclusion_str = str(conclusion_value).strip()
    # 如果包含"-"，取"-"之前的部分
    if "-" in conclusion_str:
        return conclusion_str.split("-")[0]
    # 如果没有"-"，可能是简写形式，直接返回
    return conclusion_str


def split_excel_by_description_with_images(
    input_path: str,
    output_path: str,
    desc_col_name: str = "开发结论"
):
    wb = load_workbook(input_path)
    src_ws = wb.worksheets[0]

    headers = [cell.value for cell in src_ws[1]]
    if desc_col_name not in headers:
        raise ValueError(f"未找到列：{desc_col_name}")

    desc_col_idx = headers.index(desc_col_name) + 1
    
    # 找到"开发结论说明"列的索引（如果存在）
    conclusion_desc_col_idx = None
    conclusion_desc_col_name = "开发结论说明"
    if conclusion_desc_col_name in headers:
        conclusion_desc_col_idx = headers.index(conclusion_desc_col_name) + 1
    
    # 找到"图片"列的索引（如果存在）
    img_col_idx = None
    if "图片" in headers:
        img_col_idx = headers.index("图片") + 1
    
    # 需要设置文本换行的列
    wrap_text_columns = ["商品链接", "产品标题", "主题", "核心词周期", "原因", "商品潜力说明"]
    wrap_text_col_indices = {}
    for col_name in wrap_text_columns:
        if col_name in headers:
            wrap_text_col_indices[headers.index(col_name) + 1] = col_name
    
    # 创建文本换行的对齐方式
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    # 记录列宽
    src_col_widths = {
        idx: src_ws.column_dimensions[get_column_letter(idx)].width
        for idx in range(1, len(headers) + 1)
    }

    # 记录行高
    src_row_heights = {
        row: src_ws.row_dimensions[row].height
        for row in range(1, src_ws.max_row + 1)
    }

    # 行 → 图片（预先读取所有图片数据到内存）
    row_images = {}
    for img in src_ws._images:
        row = img.anchor._from.row + 1
        try:
            # 立即读取图片数据到内存，避免文件流关闭的问题
            img_data = img._data()
            # 确保数据被完全读取到内存（转换为bytes）
            if hasattr(img_data, 'read'):
                # 如果是文件流，读取所有数据
                img_bytes = img_data.read()
                img_data.close()  # 关闭原始流
            else:
                # 如果已经是bytes，直接使用
                img_bytes = img_data
            # 将字节数据存储为元组：(图片字节数据, 列索引, 宽度, 高度)
            col_idx = img.anchor._from.col + 1
            row_images.setdefault(row, []).append((img_bytes, col_idx, img.width, img.height))
        except Exception as e:
            print(f"读取图片数据失败（行 {row}）: {e}")

    # 三层分组：
    # 第一层：按照主要类别分组（用于创建sheet）
    # 第二层：按照开发结论值分组
    # 第三层：按照开发结论说明值分组 - {主类别: {开发结论: {开发结论说明: [行索引列表]}}}
    category_groups = {}
    for row_idx in range(2, src_ws.max_row + 1):
        conclusion_value = src_ws.cell(row=row_idx, column=desc_col_idx).value
        conclusion_str = str(conclusion_value).strip() if conclusion_value else "未分类"
        
        # 获取开发结论说明值
        if conclusion_desc_col_idx:
            conclusion_desc_value = src_ws.cell(row=row_idx, column=conclusion_desc_col_idx).value
            conclusion_desc_str = str(conclusion_desc_value).strip() if conclusion_desc_value else "未分类说明"
        else:
            conclusion_desc_str = "未分类说明"
        
        main_category = extract_main_category(conclusion_str)
        
        if main_category not in category_groups:
            category_groups[main_category] = {}
        
        if conclusion_str not in category_groups[main_category]:
            category_groups[main_category][conclusion_str] = {}
        
        if conclusion_desc_str not in category_groups[main_category][conclusion_str]:
            category_groups[main_category][conclusion_str][conclusion_desc_str] = []
        
        category_groups[main_category][conclusion_str][conclusion_desc_str].append(row_idx)

    new_wb = Workbook()
    ws_original = new_wb.active
    ws_original.title = "源数据"
    
    # 定义sheet创建顺序（主要类别）
    sheet_order = ["源数据", "开发", "追踪", "待定", "不开发"]

    # 复制函数，用于复制一行数据到目标工作表
    def copy_row_to_sheet(src_ws, target_ws, old_row, new_row_idx, headers, wrap_text_col_indices, wrap_alignment, src_row_heights, row_images, src_col_widths, img_col_idx):
        # 写单元格
        for col_idx in range(1, len(headers) + 1):
            cell = target_ws.cell(
                row=new_row_idx,
                column=col_idx,
                value=src_ws.cell(row=old_row, column=col_idx).value
            )
            # 如果是需要换行的列，设置对齐方式为换行
            if col_idx in wrap_text_col_indices:
                cell.alignment = wrap_alignment

        # 复制行高
        if src_row_heights.get(old_row):
            target_ws.row_dimensions[new_row_idx].height = src_row_heights[old_row]

        # 复制图片（核心修复）
        if old_row in row_images:
            for img_data_tuple in row_images[old_row]:
                try:
                    img_bytes, col_idx, original_width, original_height = img_data_tuple
                    # 从内存中的字节数据创建新的图片
                    stream = BytesIO(img_bytes)
                    new_img = XLImage(stream)

                    col_letter = get_column_letter(col_idx)

                    # 计算目标像素尺寸
                    col_width = src_col_widths.get(col_idx)
                    row_height = src_row_heights.get(old_row)

                    max_w = excel_colwidth_to_pixels(col_width)
                    max_h = excel_rowheight_to_pixels(row_height)
                    
                    # 如果当前图片是"图片"列中的内容，则高度只取一半
                    if img_col_idx is not None and col_idx == img_col_idx:
                        max_h = max_h // 2

                    # 限制尺寸（关键）
                    new_img.width = min(new_img.width, max_w)
                    new_img.height = min(new_img.height, max_h)

                    target_ws.add_image(new_img, f"{col_letter}{new_row_idx}")

                except Exception as e:
                    print(f"图片复制失败（行 {old_row}）: {e}")

    # 1. 第一个sheet：复制所有原始数据
    # 写表头
    for col_idx, header in enumerate(headers, 1):
        ws_original.cell(row=1, column=col_idx, value=header)

    # 复制列宽
    for col_idx, width in src_col_widths.items():
        if width:
            ws_original.column_dimensions[get_column_letter(col_idx)].width = width

    original_row_idx = 2
    for old_row in range(2, src_ws.max_row + 1):
        copy_row_to_sheet(
            src_ws, ws_original, old_row, original_row_idx,
            headers, wrap_text_col_indices, wrap_alignment,
            src_row_heights, row_images, src_col_widths, img_col_idx
        )
        original_row_idx += 1

    # 定义说明内容的样式
    section_title_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
    section_title_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_font = Font(name='Microsoft YaHei', size=10, bold=True)
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 2. 其他sheet：按照主要类别创建，每个sheet内按开发结论说明值分组显示
    for main_category in sheet_order[1:]:  # 跳过"源数据"
        if main_category not in category_groups:
            continue  # 如果该主要类别没有数据，跳过
        
        # 获取该主要类别下的所有开发结论和数据
        conclusions = category_groups[main_category]
        if not conclusions:
            continue
        
        ws_new = new_wb.create_sheet(main_category[:31])
        
        # 复制列宽
        for col_idx, width in src_col_widths.items():
            if width:
                ws_new.column_dimensions[get_column_letter(col_idx)].width = width
        
        current_row = 1
        
        # 收集所有开发结论说明值，用于按说明分组
        # 因为同一个说明可能出现在不同的开发结论下，我们需要按说明分组
        all_conclusion_descriptions = {}
        for conclusion_str, conclusion_descriptions in conclusions.items():
            for desc_str, rows in conclusion_descriptions.items():
                if desc_str not in all_conclusion_descriptions:
                    all_conclusion_descriptions[desc_str] = []
                all_conclusion_descriptions[desc_str].extend(rows)
        
        # 按照开发结论说明顺序处理每个部分
        sorted_descriptions = sorted(all_conclusion_descriptions.keys())
        
        for desc_idx, conclusion_desc_str in enumerate(sorted_descriptions):
            rows = all_conclusion_descriptions[conclusion_desc_str]
            if not rows:
                continue
            
            # 1. 添加开发结论说明行（显示开发结论说明的值）
            section_title_cell = ws_new.cell(row=current_row, column=1, value=conclusion_desc_str)
            section_title_cell.font = section_title_font
            section_title_cell.fill = section_title_fill
            section_title_cell.alignment = Alignment(horizontal='left', vertical='center')
            section_title_cell.border = border_style
            
            # 合并单元格，让说明跨越多列
            max_col = min(len(headers), 10)
            ws_new.merge_cells(start_row=current_row, start_column=1, 
                              end_row=current_row, end_column=max_col)
            ws_new.row_dimensions[current_row].height = 22
            current_row += 1
            
            # 2. 写表头行
            for col_idx, header in enumerate(headers, 1):
                header_cell = ws_new.cell(row=current_row, column=col_idx, value=header)
                header_cell.font = header_font
                header_cell.border = border_style
            ws_new.row_dimensions[current_row].height = 20
            current_row += 1
            
            # 3. 复制数据行
            for old_row in rows:
                copy_row_to_sheet(
                    src_ws, ws_new, old_row, current_row,
                    headers, wrap_text_col_indices, wrap_alignment,
                    src_row_heights, row_images, src_col_widths, img_col_idx
                )
                current_row += 1
            
            # 4. 在部分之间添加空行（最后一个部分不添加）
            if desc_idx < len(sorted_descriptions) - 1:
                current_row += 2  # 空出2行作为分隔

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    new_wb.save(output_path)

    print(f"✅ 已完成拆分并保持图片大小：{output_path}")


if __name__ == '__main__':
    split_excel_by_description_with_images(input_path='result/bs/流量周期分析结果_20260109.xlsx',
                                           output_path='result/bs/潜在价值结果/分类后_潜在价值_20260109.xlsx')