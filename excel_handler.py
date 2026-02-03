import io
from typing import Dict, Optional
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from PIL import Image
import io

def bytes_to_png_bytes(image_bytes: bytes) -> io.BytesIO:
    """把任意图片 bytes 转成 PNG bytes（兼容 webp/jpg/png/gif 等）"""
    bio_in = io.BytesIO(image_bytes)
    with Image.open(bio_in) as im:
        # 统一转 RGBA，避免透明/调色板模式导致异常
        im = im.convert("RGBA")
        bio_out = io.BytesIO()
        im.save(bio_out, format="PNG")
        bio_out.seek(0)
        return bio_out



def insert_traffic_cycle_images(output_path: str, traffic_cycle_images: Dict[int, Optional[bytes]], df_index_mapping: list):
    """插入流量周期图到Excel
    
    Args:
        output_path: Excel文件路径
        traffic_cycle_images: 图片字典，key为原始DataFrame索引
        df_index_mapping: DataFrame索引列表，用于映射到Excel行号
    """
    if not traffic_cycle_images:
        return
    
    try:
        wb = load_workbook(output_path)
        ws = wb.active

        # 找到"核心词周期图"列的位置
        header_row = 1
        chart_col_idx = None
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value == "核心词周期图":
                chart_col_idx = col_idx
                break

        # 如果列不存在，创建它（在最后一列之后）
        if not chart_col_idx:
            chart_col_idx = ws.max_column + 1
            ws.cell(header_row, chart_col_idx, "核心词周期图")
            print(f'已创建"核心词周期图"列（第{chart_col_idx}列）')

        # 现在 chart_col_idx 一定存在，插入图片
        # 将图片放到"核心词周期图"列本身的位置
        if chart_col_idx:
            target_col_idx = chart_col_idx
            col_letter = get_column_letter(target_col_idx)
            # 确保目标列存在（若越界会自动扩展），并加宽以容纳图片
            if ws.column_dimensions[col_letter].width is None or ws.column_dimensions[col_letter].width < 60:
                ws.column_dimensions[col_letter].width = 60
            inserted_count = 0

            # 遍历所有行，插入图片
            # 注意：DataFrame保存后，行号是连续的（0, 1, 2, ...），对应Excel的第2, 3, 4行
            for df_idx, original_idx in enumerate(df_index_mapping):
                excel_row = df_idx + 2  # Excel行号 = DataFrame索引 + 2（表头1行 + 1）

                if original_idx in traffic_cycle_images and traffic_cycle_images[original_idx]:
                    try:
                        img_data = traffic_cycle_images[original_idx]
                        img = XLImage(io.BytesIO(img_data))

                        # 设置图片尺寸（像素）
                        img.width = 400
                        img.height = 200

                        # 设置列宽以适应图片
                        current_width = ws.column_dimensions[col_letter].width
                        if current_width is None or current_width < 60:
                            ws.column_dimensions[col_letter].width = 60

                        # 设置行高以适应图片
                        current_height = ws.row_dimensions[excel_row].height
                        if current_height is None or current_height < 150:
                            ws.row_dimensions[excel_row].height = 150

                        # 直接锚定到单元格
                        ws.add_image(img, f"{col_letter}{excel_row}")

                        inserted_count += 1
                    except Exception as e:
                        print(f"插入第 {excel_row} 行（DataFrame索引{df_idx}）流量周期图时出错: {e}")
                        import traceback
                        traceback.print_exc()

            wb.save(output_path)
            print(f'成功插入 {inserted_count} 张流量周期图到 {output_path}')
        else:
            print(f'警告: 未找到"核心词周期图"列，跳过图片插入')
    except Exception as e:
        print(f'插入流量周期图时出错: {e}')
        import traceback
        traceback.print_exc()


def insert_sales_trend_images(output_path: str, sales_trend_images: Dict[int, Optional[bytes]], df_index_mapping: list):
    """插入销量趋势图到Excel
    
    Args:
        output_path: Excel文件路径
        sales_trend_images: 图片字典，key为原始DataFrame索引
        df_index_mapping: DataFrame索引列表，用于映射到Excel行号
    """
    if not sales_trend_images:
        return
    
    try:
        wb = load_workbook(output_path)
        ws = wb.active

        # 找到"销量趋势图"列的位置
        header_row = 1
        sales_chart_col_idx = None
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value == "销量趋势图":
                sales_chart_col_idx = col_idx
                break

        # 如果列不存在，创建它（在最后一列之后）
        if not sales_chart_col_idx:
            sales_chart_col_idx = ws.max_column + 1
            ws.cell(header_row, sales_chart_col_idx, "销量趋势图")
            print(f'已创建"销量趋势图"列（第{sales_chart_col_idx}列）')

        # 现在 sales_chart_col_idx 一定存在，插入图片
        # 将图片放到"销量趋势图"列本身的位置
        if sales_chart_col_idx:
            target_col_idx = sales_chart_col_idx
            col_letter = get_column_letter(target_col_idx)
            # 确保目标列存在（若越界会自动扩展），并加宽以容纳图片
            if ws.column_dimensions[col_letter].width is None or ws.column_dimensions[col_letter].width < 50:
                ws.column_dimensions[col_letter].width = 50
            inserted_count = 0

            # 遍历所有行，插入图片
            for df_idx, original_idx in enumerate(df_index_mapping):
                excel_row = df_idx + 2  # Excel行号 = DataFrame索引 + 2（表头1行 + 1）

                if original_idx in sales_trend_images and sales_trend_images[original_idx]:
                    try:
                        img_data = sales_trend_images[original_idx]
                        img = XLImage(io.BytesIO(img_data))

                        # 设置图片尺寸（像素）
                        img.width = 350
                        img.height = 200

                        # 设置列宽以适应图片
                        current_width = ws.column_dimensions[col_letter].width
                        if current_width is None or current_width < 50:
                            ws.column_dimensions[col_letter].width = 50

                        # 设置行高以适应图片
                        current_height = ws.row_dimensions[excel_row].height
                        if current_height is None or current_height < 150:
                            ws.row_dimensions[excel_row].height = 150

                        # 直接锚定到单元格
                        ws.add_image(img, f"{col_letter}{excel_row}")

                        inserted_count += 1
                    except Exception as e:
                        print(f"插入第 {excel_row} 行（DataFrame索引{df_idx}）销量趋势图时出错: {e}")
                        import traceback
                        traceback.print_exc()

            wb.save(output_path)
            print(f'成功插入 {inserted_count} 张销量趋势图到 {output_path}')
        else:
            print(f'警告: 未找到"销量趋势图"列，跳过图片插入')
    except Exception as e:
        print(f'插入销量趋势图时出错: {e}')
        import traceback
        traceback.print_exc()


def insert_price_trend_images(output_path: str, price_trend_images: Dict[int, Optional[bytes]], df_index_mapping: list):
    """插入价格趋势图到Excel
    
    Args:
        output_path: Excel文件路径
        price_trend_images: 图片字典，key为原始DataFrame索引
        df_index_mapping: DataFrame索引列表，用于映射到Excel行号
    """
    print(f'准备插入价格趋势图，price_trend_images字典中有 {len(price_trend_images)} 个条目')
    try:
        wb = load_workbook(output_path)
        ws = wb.active

        # 找到"价格趋势图"列的位置
        header_row = 1
        price_chart_col_idx = None
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value == "价格趋势图":
                price_chart_col_idx = col_idx
                print(f'找到"价格趋势图"列，位于第{price_chart_col_idx}列')
                break

        # 如果列不存在，创建它（在最后一列之后）
        if not price_chart_col_idx:
            price_chart_col_idx = ws.max_column + 1
            ws.cell(header_row, price_chart_col_idx, "价格趋势图")
            print(f'已创建"价格趋势图"列（第{price_chart_col_idx}列）')

        # 现在 price_chart_col_idx 一定存在，插入图片
        # 将图片放到"价格趋势图"列本身的位置
        if price_chart_col_idx:
            target_col_idx = price_chart_col_idx
            col_letter = get_column_letter(target_col_idx)
            # 确保目标列存在（若越界会自动扩展），并加宽以容纳图片
            if ws.column_dimensions[col_letter].width is None or ws.column_dimensions[col_letter].width < 50:
                ws.column_dimensions[col_letter].width = 50
            inserted_count = 0

            # 遍历所有行，插入图片
            for df_idx, original_idx in enumerate(df_index_mapping):
                excel_row = df_idx + 2  # Excel行号 = DataFrame索引 + 2（表头1行 + 1）

                if original_idx in price_trend_images and price_trend_images[original_idx]:
                    try:
                        img_data = price_trend_images[original_idx]
                        img = XLImage(io.BytesIO(img_data))

                        # 设置图片尺寸（像素）
                        img.width = 350
                        img.height = 200

                        # 设置列宽以适应图片
                        current_width = ws.column_dimensions[col_letter].width
                        if current_width is None or current_width < 50:
                            ws.column_dimensions[col_letter].width = 50

                        # 设置行高以适应图片
                        current_height = ws.row_dimensions[excel_row].height
                        if current_height is None or current_height < 150:
                            ws.row_dimensions[excel_row].height = 150

                        # 直接锚定到单元格
                        ws.add_image(img, f"{col_letter}{excel_row}")

                        inserted_count += 1
                        print(f"  ✓ 插入价格趋势图: Excel行{excel_row}, DataFrame索引{df_idx}, 原始索引{original_idx}")
                    except Exception as e:
                        print(f"插入第 {excel_row} 行（DataFrame索引{df_idx}）价格趋势图时出错: {e}")
                        import traceback
                        traceback.print_exc()
                else:
                    if original_idx not in price_trend_images:
                        print(f"  - 跳过行{excel_row}（原始索引{original_idx}）: 价格趋势图数据不存在")
                    elif not price_trend_images[original_idx]:
                        print(f"  - 跳过行{excel_row}（原始索引{original_idx}）: 价格趋势图数据为空")

            wb.save(output_path)
            print(f'成功插入 {inserted_count} 张价格趋势图到 {output_path}')
        else:
            print(f'警告: 未找到"价格趋势图"列，跳过图片插入')
    except Exception as e:
        print(f'插入价格趋势图时出错: {e}')
        import traceback
        traceback.print_exc()


def insert_product_images(output_path: str):
    """根据图片链接插入产品图片（在新列"图片"中展示）"""
    try:
        wb = load_workbook(output_path)
        ws = wb.active

        # 找到"图片链接"列位置
        header_row = 1
        link_col_idx = None
        img_col_idx = None
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value == "图片链接":
                link_col_idx = col_idx
            if cell.value == "图片":
                img_col_idx = col_idx

        if link_col_idx is None:
            print('警告: 未找到"图片链接"列，跳过图片插入')
        else:
            # 如果"图片"列不存在，插入到"图片链接"列的位置
            # 注意：插入新列后，"图片链接"列会移动到下一列
            if img_col_idx is None:
                img_col_idx = link_col_idx
                ws.insert_cols(img_col_idx)  # 插入一列
                ws.cell(header_row, img_col_idx, "图片")
                print(f'已创建"图片"列（第{img_col_idx}列，位于"图片链接"列位置）')
                # 插入新列后，"图片链接"列移动到下一列
                link_col_idx = link_col_idx + 1
                
                # 立即清除新插入列的所有单元格值和超链接（防止复制了原列的内容）
                # 注意：insert_cols可能会复制原列内容，必须立即清除
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row, img_col_idx)
                    # 强制清除值（无论是什么值都清除，包括空字符串）
                    cell.value = None
                    # 清除超链接
                    if cell.hyperlink is not None:
                        cell.hyperlink = None

            col_letter = get_column_letter(img_col_idx)
            
            # 清除整个"图片"列的所有单元格值和超链接（确保没有URL文本和超链接）
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, img_col_idx)
                # 强制清除值（无论是什么值都清除）
                cell.value = None
                # 清除超链接
                if cell.hyperlink is not None:
                    cell.hyperlink = None
            
            inserted_img_count = 0
            session = requests.Session()

            for row in range(2, ws.max_row + 1):
                url = ws.cell(row, link_col_idx).value

                if not url:
                    continue

                # 重试机制：最多重试3次
                max_retries = 3
                success = False
                for attempt in range(max_retries):
                    try:
                        resp = session.get(str(url), timeout=10)
                        resp.raise_for_status()
                        # img_bytes = io.BytesIO(resp.content)  # 碰见.webp保存的问题
                        # img = XLImage(img_bytes)
                        png_io = bytes_to_png_bytes(resp.content)  # ✅ 关键：转 PNG
                        img = XLImage(png_io)
                        img.width = 75
                        img.height = 75

                        # 设置列宽和行高
                        current_width = ws.column_dimensions[col_letter].width
                        if current_width is None or current_width < 14:
                            ws.column_dimensions[col_letter].width = 14

                        current_height = ws.row_dimensions[row].height
                        if current_height is None or current_height < 80:
                            ws.row_dimensions[row].height = 80

                        # 直接锚定到对应单元格
                        ws.add_image(img, f"{col_letter}{row}")
                        
                        # 插入图片后，立即清除单元格的值和超链接（确保没有URL文本和超链接）
                        cell = ws.cell(row, img_col_idx)
                        # 强制清除值（无论是什么值都清除）
                        cell.value = None
                        # 清除超链接
                        if cell.hyperlink is not None:
                            cell.hyperlink = None
                        
                        inserted_img_count += 1
                        success = True
                        if attempt > 0:
                            print(f"  ✓ 插入图片（重试{attempt}次后成功）: Excel行{row}, url={url}")
                        else:
                            print(f"  ✓ 插入图片: Excel行{row}, url={url}")
                        break  # 成功，退出重试循环
                    except Exception as e:
                        if attempt < max_retries - 1:
                            # 不是最后一次尝试，继续重试
                            print(f"  ⚠ 插入图片失败（尝试 {attempt + 1}/{max_retries}）: Excel行{row}, url={url}, 错误: {e}")
                            continue
                        else:
                            # 最后一次尝试也失败了
                            print(f"  ✗ 插入图片失败（已重试{max_retries}次）: Excel行{row}, url={url}, 错误: {e}")
                            break

            # 保存前，再次清除图片列所有单元格的值和超链接（确保没有URL残留）
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, img_col_idx)
                # 强制清除值（无论是什么值都清除）
                cell.value = None
                # 清除超链接
                if cell.hyperlink is not None:
                    cell.hyperlink = None

            wb.save(output_path)
            print(f'成功插入 {inserted_img_count} 张产品图片到 {output_path}')
            
            # 保存后重新加载文件，最后一次清除图片列的所有值和超链接（彻底解决URL残留问题）
            wb_final = load_workbook(output_path)
            ws_final = wb_final.active
            
            # 重新找到图片列的位置
            img_col_idx_final = None
            for col_idx, cell in enumerate(ws_final[header_row], 1):
                if cell.value == "图片":
                    img_col_idx_final = col_idx
                    break
            
            if img_col_idx_final:
                for row in range(2, ws_final.max_row + 1):
                    cell = ws_final.cell(row, img_col_idx_final)
                    # 强制清除值（无论是什么值都清除，包括URL字符串）
                    cell.value = None
                    # 清除超链接
                    if cell.hyperlink is not None:
                        cell.hyperlink = None
                
                wb_final.save(output_path)
                print(f'✅ 已彻底清除图片列中的所有URL文本和超链接')

            # 调整指定列的列宽
            # 产品标题、核心词周期、核心词周期图列加宽
            target_columns = ["产品标题", "核心词周期", "核心词周期图"]
            desired_widths = {
                "产品标题": 50,
                "核心词周期": 40,
                "核心词周期图": 40
            }
            for col_idx, cell in enumerate(ws[header_row], 1):
                title = cell.value
                if title in target_columns:
                    width = desired_widths.get(title, 30)
                    col_letter = get_column_letter(col_idx)
                    current_width = ws.column_dimensions[col_letter].width
                    if current_width is None or current_width < width:
                        ws.column_dimensions[col_letter].width = width
    except Exception as e:
        print(f'插入产品图片时出错: {e}')
        import traceback
        traceback.print_exc()


def delete_column_from_excel(output_path: str, column_name: str):
    """从Excel文件中删除指定列
    
    Args:
        output_path: Excel文件路径
        column_name: 要删除的列名
    """
    try:
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 找到要删除的列位置
        header_row = 1
        col_idx = None
        for idx, cell in enumerate(ws[header_row], 1):
            if cell.value == column_name:
                col_idx = idx
                break
        
        if col_idx is None:
            print(f'警告: 未找到"{column_name}"列，跳过删除操作')
            return
        
        # 删除列
        ws.delete_cols(col_idx)
        wb.save(output_path)
        print(f'成功删除"{column_name}"列（原第{col_idx}列）')
    except Exception as e:
        print(f'删除列"{column_name}"时出错: {e}')
        import traceback
        traceback.print_exc()
